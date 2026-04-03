import os
import re
import json
import tempfile
import requests
import openpyxl
from bs4 import BeautifulSoup
from flask import Flask, render_template, request

app = Flask(__name__)

COLORS = [
    '#e63946', '#457b9d', '#2a9d8f', '#e9c46a', '#f4a261',
    '#264653', '#a8dadc', '#6d6875', '#b5838d', '#ffb4a2'
]

# ── PARSING ────────────────────────────────────────────────────────────────────

def parse_result(result_str):
    if not result_str or result_str == '-':
        return None
    try:
        g_casa, g_ospite = result_str.split('-')
        g_casa, g_ospite = int(g_casa), int(g_ospite)
        if g_casa > g_ospite:
            return (3, 0)
        elif g_casa == g_ospite:
            return (1, 1)
        else:
            return (0, 3)
    except Exception:
        return None


def parse_excel(path):
    wb = openpyxl.load_workbook(path)

    # Trova il foglio giusto: prima cerca "Calendario", poi quello che
    # contiene "Giornata lega", infine usa il primo foglio disponibile.
    ws = None
    if 'Calendario' in wb.sheetnames:
        ws = wb['Calendario']
    else:
        for name in wb.sheetnames:
            sheet = wb[name]
            for row in sheet.iter_rows(min_row=1, max_row=20, values_only=True):
                if any(isinstance(c, str) and 'Giornata lega' in c for c in row if c):
                    ws = sheet
                    break
            if ws:
                break
        if ws is None:
            ws = wb[wb.sheetnames[0]]

    # Prova a leggere il nome lega dalle prime righe
    league_name = None
    for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
        for cell in row:
            if isinstance(cell, str) and len(cell) > 2 and 'Giornata' not in cell:
                league_name = cell.strip()
                break
        if league_name:
            break

    matchdays = {}
    rows = list(ws.iter_rows(min_row=4, values_only=True))

    i = 0
    while i < len(rows):
        row = rows[i]
        left_header = row[0]
        right_header = row[6] if len(row) > 6 else None

        if isinstance(left_header, str) and 'Giornata lega' in left_header:
            left_num = int(left_header.split('ª')[0])
            right_num = None
            if isinstance(right_header, str) and 'Giornata lega' in right_header:
                right_num = int(right_header.split('ª')[0])

            if left_num not in matchdays:
                matchdays[left_num] = []
            if right_num and right_num not in matchdays:
                matchdays[right_num] = []

            for j in range(1, 6):
                if i + j >= len(rows):
                    break
                match_row = rows[i + j]

                casa_l = match_row[0]
                ospite_l = match_row[3] if len(match_row) > 3 else None
                res_l = match_row[4] if len(match_row) > 4 else None
                pts = parse_result(res_l)
                if pts and casa_l and ospite_l:
                    matchdays[left_num].append((str(casa_l), str(ospite_l), pts[0], pts[1]))

                if right_num and len(match_row) > 10:
                    casa_r = match_row[6]
                    ospite_r = match_row[9]
                    res_r = match_row[10]
                    pts2 = parse_result(res_r)
                    if pts2 and casa_r and ospite_r:
                        matchdays[right_num].append((str(casa_r), str(ospite_r), pts2[0], pts2[1]))

            i += 6
        else:
            i += 1

    return matchdays, league_name


def compute_standings(matchdays):
    teams = set()
    for matches in matchdays.values():
        for casa, ospite, _, _ in matches:
            teams.add(casa)
            teams.add(ospite)
    teams = sorted(teams)

    current_pts = {t: 0 for t in teams}
    standings_over_time = []
    giornate_labels = []

    for giornata_num in sorted(matchdays.keys()):
        matches = matchdays[giornata_num]
        if not matches:
            continue
        for casa, ospite, pts_c, pts_o in matches:
            current_pts[casa] += pts_c
            current_pts[ospite] += pts_o
        standings_over_time.append(dict(current_pts))
        giornate_labels.append(f'G{giornata_num}')

    return teams, standings_over_time, giornate_labels


# ── LOGO FETCHING (fantacalcio.it + fantagazzetta.com) ────────────────────────

_BROWSER_HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) '
                  'AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36',
    'Accept-Language': 'it-IT,it;q=0.9,en;q=0.8',
}


def _extract_logos_from_json(data, depth=0):
    """Cerca ricorsivamente coppie {name, logo} in qualsiasi JSON."""
    if depth > 10:
        return {}
    logos = {}
    if isinstance(data, dict):
        name = (data.get('name') or data.get('teamName') or
                data.get('team_name') or data.get('squadra') or '')
        logo = (data.get('logo') or data.get('logoUrl') or data.get('logo_url') or
                data.get('image') or data.get('imageUrl') or data.get('image_url') or
                data.get('team_logo') or data.get('crest') or data.get('picture') or '')
        if name and logo and isinstance(name, str) and isinstance(logo, str):
            logos[name.strip()] = logo.strip()
        for v in data.values():
            logos.update(_extract_logos_from_json(v, depth + 1))
    elif isinstance(data, list):
        for item in data:
            logos.update(_extract_logos_from_json(item, depth + 1))
    return logos


def _try_api_endpoints(endpoints):
    """Prova una lista di endpoint JSON, ritorna il primo dict non vuoto."""
    headers = {**_BROWSER_HEADERS, 'Accept': 'application/json'}
    for url in endpoints:
        try:
            resp = requests.get(url, headers=headers, timeout=8)
            if resp.status_code != 200:
                continue
            data = resp.json()
            logos = _extract_logos_from_json(data)
            if logos:
                return logos
        except Exception:
            continue
    return {}


def _scrape_page_for_logos(url):
    """
    Scarica la pagina HTML e cerca:
    1. JSON embedded in <script> (Nuxt __NUXT__, Next __NEXT_DATA__, ecc.)
    2. Tag <img> con attributi che suggeriscono loghi squadra
    """
    logos = {}
    try:
        resp = requests.get(url, headers={**_BROWSER_HEADERS, 'Accept': 'text/html'}, timeout=12)
        if resp.status_code != 200:
            return logos
        html = resp.text
        soup = BeautifulSoup(html, 'html.parser')

        # 1. Script tag con JSON embedded
        for script in soup.find_all('script'):
            content = script.string or ''
            # Nuxt / Next / generici pattern
            for pattern in [
                r'window\.__NUXT__\s*=\s*(\{.+)',
                r'<script[^>]+id=["\']__NUXT_DATA__["\'][^>]*>(.+)',
                r'window\.__INITIAL_STATE__\s*=\s*(\{.+)',
                r'window\.__APP_STATE__\s*=\s*(\{.+)',
            ]:
                m = re.search(pattern, content, re.DOTALL)
                if m:
                    # Tenta di parsare raccogliendo JSON valido
                    raw = m.group(1).rstrip(';')
                    try:
                        data = json.loads(raw)
                        found = _extract_logos_from_json(data)
                        logos.update(found)
                    except Exception:
                        pass

            # Script type="application/json"
            if script.get('type') == 'application/json':
                try:
                    data = json.loads(content)
                    logos.update(_extract_logos_from_json(data))
                except Exception:
                    pass

        if logos:
            return logos

        # 2. Tag <img> con src che contiene parole chiave legate a loghi
        base_url = '/'.join(url.split('/')[:3])
        for img in soup.find_all('img'):
            src = img.get('src', '') or img.get('data-src', '')
            alt = img.get('alt', '').strip()
            if not src or not alt:
                continue
            if any(kw in src.lower() for kw in ['logo', 'team', 'squadra', 'crest', 'shield']):
                if src.startswith('/'):
                    src = base_url + src
                logos[alt] = src

    except Exception:
        pass
    return logos


def fetch_logos(league_input):
    """
    Recupera i loghi dalla lega. Supporta:
    - leghe.fantacalcio.it/{slug}/...   (slug testuale, no ID numerico)
    - leghe.fantagazzetta.com/.../{id}  (ID numerico)
    """
    if not league_input:
        return {}
    league_input = league_input.strip()

    # ── fantacalcio.it ──────────────────────────────────────────────────────
    if 'fantacalcio.it' in league_input:
        m = re.search(r'leghe\.fantacalcio\.it/([^/?#]+)', league_input)
        if not m:
            return {}
        slug = m.group(1)
        base = f'https://leghe.fantacalcio.it/{slug}'

        # Prima tenta API REST con slug
        logos = _try_api_endpoints([
            f'https://leghe.fantacalcio.it/api/v1/leagues/{slug}/participants',
            f'https://leghe.fantacalcio.it/api/v1/leagues/{slug}/ranking',
            f'https://leghe.fantacalcio.it/api/v1/leagues/{slug}/teams',
            f'https://leghe.fantacalcio.it/api/leagues/{slug}/participants',
            f'https://leghe.fantacalcio.it/api/{slug}/teams',
        ])
        if logos:
            return logos

        # Poi scraping delle pagine più ricche di dati
        for page in [f'{base}/classifica', f'{base}/squadre', f'{base}/calendario', league_input]:
            logos = _scrape_page_for_logos(page)
            if logos:
                return logos
        return {}

    # ── fantagazzetta.com (ID numerico) ─────────────────────────────────────
    if 'fantagazzetta' in league_input or league_input.isdigit():
        league_id = league_input if league_input.isdigit() else ''
        if not league_id:
            m = re.search(r'(\d{4,})', league_input)
            if m:
                league_id = m.group(1)
        if not league_id:
            return {}
        return _try_api_endpoints([
            f'https://leghe.fantagazzetta.com/api/v1/leagues/{league_id}/standings',
            f'https://leghe.fantagazzetta.com/api/v1/leagues/{league_id}/ranking',
            f'https://leghe.fantagazzetta.com/api/v1/leagues/{league_id}/participants',
            f'https://leghe.fantagazzetta.com/api/v1/leagues/{league_id}/teams',
        ])

    return {}


# ── ROUTES ────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload():
    if 'excel' not in request.files or request.files['excel'].filename == '':
        return render_template('index.html', error='Seleziona un file Excel.')

    file = request.files['excel']
    league_input = request.form.get('league_url', '').strip()

    # Salva temporaneamente il file
    suffix = os.path.splitext(file.filename)[1] or '.xlsx'
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        file.save(tmp.name)
        tmp_path = tmp.name

    try:
        matchdays, league_name = parse_excel(tmp_path)
    except Exception as e:
        os.unlink(tmp_path)
        return render_template('index.html', error=f'Errore nel parsing: {e}')
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)

    if not matchdays:
        return render_template('index.html', error='Nessuna giornata trovata nel file.')

    teams, standings_over_time, giornate_labels = compute_standings(matchdays)

    # Nome lega: sempre dal filename (più affidabile del contenuto Excel)
    base = os.path.splitext(file.filename)[0]
    league_name = re.sub(r'^[Cc]alendario[_\s]*', '', base).replace('_', ' ').strip()
    if not league_name:
        league_name = base.replace('_', ' ').strip()

    # Colori squadre
    color_map = {team: COLORS[i % len(COLORS)] for i, team in enumerate(sorted(teams))}

    # Loghi Fantagazzetta
    logos = {}
    logo_error = None
    if league_input:
        logos = fetch_logos(league_input)
        if not logos:
            logo_error = 'Loghi non trovati. Verifica URL/ID della lega.'

    dashboard_data = {
        'league_name': league_name,
        'teams': teams,
        'standings': standings_over_time,
        'giornate': giornate_labels,
        'colors': color_map,
        'logos': logos,
    }

    return render_template(
        'dashboard.html',
        data=json.dumps(dashboard_data, ensure_ascii=False),
        league_name=league_name,
        logo_error=logo_error,
        has_logos=bool(logos),
    )


if __name__ == '__main__':
    app.run(debug=True, port=5050)
